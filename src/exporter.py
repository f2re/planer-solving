import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from typing import List, Dict, Tuple, Any

def export_to_excel(transformed_data: Dict[str, Any], teachers_config: List[Dict], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Лист1'
    
    dates = transformed_data['dates']
    grid = transformed_data['grid']
    
    # Fills for different lesson types
    fills = {
        'Л': PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid'), # LightYellow
        'П': PatternFill(start_color='E0FFFF', end_color='E0FFFF', fill_type='solid'), # LightCyan
        'С': PatternFill(start_color='E0FFFF', end_color='E0FFFF', fill_type='solid'),
        'ЛР': PatternFill(start_color='E0FFE0', end_color='E0FFE0', fill_type='solid'), # LightGreen
        'Экз': PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid'), # LightRed
    }
    
    # 1. Header Styles
    header_font = Font(name='Times New Roman', size=8, bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_side = Side(style='thin')
    medium_side = Side(style='medium')
    
    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5 + len(dates))
    
    # Form dynamic title
    semester_part = transformed_data.get('semester_info', 'На семестр')
    # Clean up "Расписание учебных занятий" prefix if present
    if "расписание учебных занятий" in semester_part.lower():
        semester_part = semester_part.lower().replace("расписание учебных занятий", "").strip()
        semester_part = semester_part.capitalize()
    
    if "на " not in semester_part.lower() and semester_part:
        semester_part = f"На {semester_part.lower()}"
    
    year_part = transformed_data.get('year_info', '')
    title_text = f"{semester_part} {year_part}".strip()
    
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = header_font
    title_cell.alignment = Alignment(horizontal='left', vertical='center')

    # Row 2-4: Column Names and Dates
    cols = ['№ п/п', 'Должность', 'Звание', 'Фамилия, имя, отчество', 'Месяц']
    for i, col_name in enumerate(cols):
        cell = ws.cell(row=2, column=i+1, value=col_name)
        cell.font = header_font
        cell.alignment = header_alignment
        if col_name != 'Месяц':
             ws.merge_cells(start_row=2, start_column=i+1, end_row=4, end_column=i+1)

    ws.cell(row=3, column=5, value='№ недели').font = header_font
    ws.cell(row=3, column=5, value='№ недели').alignment = header_alignment
    ws.cell(row=4, column=5, value='№ часа').font = header_font
    ws.cell(row=4, column=5, value='№ часа').alignment = header_alignment
    
    # Set widths exactly as in sample
    ws.column_dimensions['A'].width = 8.57
    ws.column_dimensions['B'].width = 16.43
    ws.column_dimensions['C'].width = 15.86
    ws.column_dimensions['D'].width = 24.71
    ws.column_dimensions['E'].width = 5.71

    # Fill Dates and Months
    current_col = 6
    last_month = None
    month_start_col = 6
    for i, (month, day) in enumerate(dates):
        cell = ws.cell(row=4, column=current_col, value=day)
        cell.font = header_font
        cell.alignment = header_alignment
        
        ws.column_dimensions[openpyxl.utils.get_column_letter(current_col)].width = 13.0
        
        if month != last_month:
            if last_month is not None:
                ws.merge_cells(start_row=2, start_column=month_start_col, end_row=2, end_column=current_col-1)
                m_cell = ws.cell(row=2, column=month_start_col, value=last_month.upper())
                m_cell.font = header_font
                m_cell.alignment = header_alignment
            last_month = month
            month_start_col = current_col
        current_col += 1
        
    if last_month is not None:
        ws.merge_cells(start_row=2, start_column=month_start_col, end_row=2, end_column=current_col-1)
        m_cell = ws.cell(row=2, column=month_start_col, value=last_month.upper())
        m_cell.font = header_font
        m_cell.alignment = header_alignment

    # 2. Fill Teachers
    current_row = 5
    data_font = Font(name='Calibri', size=11)
    
    for idx, t_info in enumerate(teachers_config):
        teacher_name = t_info['short_name']
        
        # Info cells
        c1 = ws.cell(row=current_row, column=1, value=idx + 1)
        c2 = ws.cell(row=current_row, column=2, value=t_info.get('position', ''))
        c3 = ws.cell(row=current_row, column=3, value=t_info.get('rank', ''))
        c4 = ws.cell(row=current_row, column=4, value=t_info['full_name'])
        
        for c in [c1, c2, c3, c4]:
            c.font = data_font
            c.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            ws.merge_cells(start_row=current_row, start_column=c.column, end_row=current_row + 3, end_column=c.column)

        for p in range(1, 5):
            pair_label = {1: '1-2', 2: '3-4', 3: '5-6', 4: '7-8'}[p]
            p_cell = ws.cell(row=current_row + p - 1, column=5, value=pair_label)
            p_cell.font = header_font
            p_cell.alignment = header_alignment
            
            for d_idx, (month, day) in enumerate(dates):
                item = grid.get((teacher_name, p, month, day))
                cell = ws.cell(row=current_row + p - 1, column=6 + d_idx)
                if item:
                    groups_str = ", ".join(item['groups'])
                    cell_val = f"{item['type']}\n{item['subject']}\n{item['room']}\n{groups_str}"
                    cell.value = cell_val
                    cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center', shrinkToFit=True)
                    cell.font = Font(size=8)
                    
                    l_type = item['type'].split('/')[0] if '/' in item['type'] else item['type']
                    if l_type in fills:
                        cell.fill = fills[l_type]
        
        current_row += 4

    # Apply borders
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    wb.save(output_path)
